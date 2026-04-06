from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()

# --- MAPPING TABLE 1 ---
ws1 = wb.active
ws1.title = "Mapping Table 1"

headers1 = ["Source File", "Source Attribute", "Mapping", "Staging Table Type", "Staging Table", "Staging Attribute"]
ws1.append(headers1)

# Format header
header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
header_font = Font(bold=True)
for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

data1 = [
    # SDR Movement
    ("wsdr.csv", "UPC", "Copy", "Relation", "stg_Movement_SDR", "UPC"),
    ("wsdr.csv", "STORE", "Copy", "Relation", "stg_Movement_SDR", "STORE"),
    ("wsdr.csv", "WEEK", "Copy", "Relation", "stg_Movement_SDR", "WEEK"),
    ("wsdr.csv", "MOVE", "Copy", "Relation", "stg_Movement_SDR", "MOVE"),
    ("wsdr.csv", "QTY", "Copy", "Relation", "stg_Movement_SDR", "QTY"),
    ("wsdr.csv", "PRICE", "Copy", "Relation", "stg_Movement_SDR", "PRICE"),
    ("wsdr.csv", "SALE", "Copy", "Relation", "stg_Movement_SDR", "SALE"),
    ("wsdr.csv", "PROFIT", "Copy", "Relation", "stg_Movement_SDR", "PROFIT"),
    ("wsdr.csv", "OK", "Copy", "Relation", "stg_Movement_SDR", "OK"),
    ("wsdr.csv", "(filename)", "Derive: extract 'SDR'", "Relation", "stg_Movement_SDR", "CATEGORY_CODE"),
    
    # CSO Movement
    ("WCSO-Done.csv", "UPC", "Copy", "Relation", "stg_Movement_CSO", "UPC"),
    ("WCSO-Done.csv", "STORE", "Copy", "Relation", "stg_Movement_CSO", "STORE"),
    ("WCSO-Done.csv", "WEEK", "Copy", "Relation", "stg_Movement_CSO", "WEEK"),
    ("WCSO-Done.csv", "MOVE", "Copy", "Relation", "stg_Movement_CSO", "MOVE"),
    ("WCSO-Done.csv", "QTY", "Copy", "Relation", "stg_Movement_CSO", "QTY"),
    ("WCSO-Done.csv", "PRICE", "Copy", "Relation", "stg_Movement_CSO", "PRICE"),
    ("WCSO-Done.csv", "SALE", "Copy", "Relation", "stg_Movement_CSO", "SALE"),
    ("WCSO-Done.csv", "PROFIT", "Copy", "Relation", "stg_Movement_CSO", "PROFIT"),
    ("WCSO-Done.csv", "OK", "Copy", "Relation", "stg_Movement_CSO", "OK"),
    ("WCSO-Done.csv", "(filename)", "Derive: extract 'CSO'", "Relation", "stg_Movement_CSO", "CATEGORY_CODE"),

    # TPA Movement
    ("WTPA_done.csv", "UPC", "Copy", "Relation", "stg_Movement_TPA", "UPC"),
    ("WTPA_done.csv", "STORE", "Copy", "Relation", "stg_Movement_TPA", "STORE"),
    ("WTPA_done.csv", "WEEK", "Copy", "Relation", "stg_Movement_TPA", "WEEK"),
    ("WTPA_done.csv", "MOVE", "Copy", "Relation", "stg_Movement_TPA", "MOVE"),
    ("WTPA_done.csv", "QTY", "Copy", "Relation", "stg_Movement_TPA", "QTY"),
    ("WTPA_done.csv", "PRICE", "Copy", "Relation", "stg_Movement_TPA", "PRICE"),
    ("WTPA_done.csv", "SALE", "Copy", "Relation", "stg_Movement_TPA", "SALE"),
    ("WTPA_done.csv", "PROFIT", "Copy", "Relation", "stg_Movement_TPA", "PROFIT"),
    ("WTPA_done.csv", "OK", "Copy", "Relation", "stg_Movement_TPA", "OK"),
    ("WTPA_done.csv", "(filename)", "Derive: extract 'TPA'", "Relation", "stg_Movement_TPA", "CATEGORY_CODE"),

    # CRA Movement
    ("Done-WCRA.csv", "UPC", "Copy", "Relation", "stg_Movement_CRA", "UPC"),
    ("Done-WCRA.csv", "STORE", "Copy", "Relation", "stg_Movement_CRA", "STORE"),
    ("Done-WCRA.csv", "WEEK", "Copy", "Relation", "stg_Movement_CRA", "WEEK"),
    ("Done-WCRA.csv", "MOVE", "Copy", "Relation", "stg_Movement_CRA", "MOVE"),
    ("Done-WCRA.csv", "QTY", "Copy", "Relation", "stg_Movement_CRA", "QTY"),
    ("Done-WCRA.csv", "PRICE", "Copy", "Relation", "stg_Movement_CRA", "PRICE"),
    ("Done-WCRA.csv", "SALE", "Copy", "Relation", "stg_Movement_CRA", "SALE"),
    ("Done-WCRA.csv", "PROFIT", "Copy", "Relation", "stg_Movement_CRA", "PROFIT"),
    ("Done-WCRA.csv", "OK", "Copy", "Relation", "stg_Movement_CRA", "OK"),
    ("Done-WCRA.csv", "(filename)", "Derive: extract 'CRA'", "Relation", "stg_Movement_CRA", "CATEGORY_CODE"),

    # UPCSDR
    ("UPCSDR.csv", "COM_CODE", "Copy", "Relation", "stg_Product_SDR", "COM_CODE"),
    ("UPCSDR.csv", "UPC", "Copy", "Relation", "stg_Product_SDR", "UPC"),
    ("UPCSDR.csv", "DESCRIP", "Copy", "Relation", "stg_Product_SDR", "DESCRIP"),
    ("UPCSDR.csv", "SIZE", "Copy", "Relation", "stg_Product_SDR", "SIZE"),
    ("UPCSDR.csv", "CASE", "Copy", "Relation", "stg_Product_SDR", "CASE_PACK"),
    ("UPCSDR.csv", "NITEM", "Copy", "Relation", "stg_Product_SDR", "NITEM"),
    ("UPCSDR.csv", "(filename)", "Derive: extract 'SDR'", "Relation", "stg_Product_SDR", "CATEGORY_CODE"),

    # UPCCSO
    ("UPCCSO.csv", "COM_CODE", "Copy", "Relation", "stg_Product_CSO", "COM_CODE"),
    ("UPCCSO.csv", "UPC", "Copy", "Relation", "stg_Product_CSO", "UPC"),
    ("UPCCSO.csv", "DESCRIP", "Copy", "Relation", "stg_Product_CSO", "DESCRIP"),
    ("UPCCSO.csv", "SIZE", "Copy", "Relation", "stg_Product_CSO", "SIZE"),
    ("UPCCSO.csv", "CASE", "Copy", "Relation", "stg_Product_CSO", "CASE_PACK"),
    ("UPCCSO.csv", "NITEM", "Copy", "Relation", "stg_Product_CSO", "NITEM"),
    ("UPCCSO.csv", "(filename)", "Derive: extract 'CSO'", "Relation", "stg_Product_CSO", "CATEGORY_CODE"),

    # UPCTPA
    ("UPCTPA.csv", "COM_CODE", "Copy", "Relation", "stg_Product_TPA", "COM_CODE"),
    ("UPCTPA.csv", "UPC", "Copy", "Relation", "stg_Product_TPA", "UPC"),
    ("UPCTPA.csv", "DESCRIP", "Copy", "Relation", "stg_Product_TPA", "DESCRIP"),
    ("UPCTPA.csv", "SIZE", "Copy", "Relation", "stg_Product_TPA", "SIZE"),
    ("UPCTPA.csv", "CASE", "Copy", "Relation", "stg_Product_TPA", "CASE_PACK"),
    ("UPCTPA.csv", "NITEM", "Copy", "Relation", "stg_Product_TPA", "NITEM"),
    ("UPCTPA.csv", "(filename)", "Derive: extract 'TPA'", "Relation", "stg_Product_TPA", "CATEGORY_CODE"),

    # UPCCRA
    ("UPCCRA.csv", "COM_CODE", "Copy", "Relation", "stg_Product_CRA", "COM_CODE"),
    ("UPCCRA.csv", "UPC", "Copy", "Relation", "stg_Product_CRA", "UPC"),
    ("UPCCRA.csv", "DESCRIP", "Copy", "Relation", "stg_Product_CRA", "DESCRIP"),
    ("UPCCRA.csv", "SIZE", "Copy", "Relation", "stg_Product_CRA", "SIZE"),
    ("UPCCRA.csv", "CASE", "Copy", "Relation", "stg_Product_CRA", "CASE_PACK"),
    ("UPCCRA.csv", "NITEM", "Copy", "Relation", "stg_Product_CRA", "NITEM"),
    ("UPCCRA.csv", "(filename)", "Derive: extract 'CRA'", "Relation", "stg_Product_CRA", "CATEGORY_CODE"),

    # DEMO
    ("DEMO.csv", "STORE", "Copy", "Relation", "stg_Store", "STORE"),
    ("DEMO.csv", "NAME", "Copy", "Relation", "stg_Store", "NAME"),
    ("DEMO.csv", "CITY", "Copy", "Relation", "stg_Store", "CITY"),
    ("DEMO.csv", "ZIP", "Copy", "Relation", "stg_Store", "ZIP"),
    ("DEMO.csv", "ZONE", "Copy", "Relation", "stg_Store", "ZONE"),
    ("DEMO.csv", "URBAN", "Copy", "Relation", "stg_Store", "URBAN"),
    ("DEMO.csv", "WEEKVOL", "Copy", "Relation", "stg_Store", "WEEKVOL"),
    ("DEMO.csv", "INCOME", "Copy", "Relation", "stg_Store", "INCOME"),
    ("DEMO.csv", "EDUC", "Copy", "Relation", "stg_Store", "EDUC"),
    ("DEMO.csv", "POVERTY", "Copy", "Relation", "stg_Store", "POVERTY"),
    ("DEMO.csv", "HSIZEAVG", "Copy", "Relation", "stg_Store", "HSIZEAVG"),
    ("DEMO.csv", "ETHNIC", "Copy", "Relation", "stg_Store", "ETHNIC"),
    ("DEMO.csv", "DENSITY", "Copy", "Relation", "stg_Store", "DENSITY"),
    ("DEMO.csv", "PRICLOW", "Copy", "Relation", "stg_Store", "PRICLOW"),
    ("DEMO.csv", "PRICMED", "Copy", "Relation", "stg_Store", "PRICMED"),
    ("DEMO.csv", "PRICHIGH", "Copy", "Relation", "stg_Store", "PRICHIGH"),
    ("DEMO.csv", "AGE9", "Copy", "Relation", "stg_Store", "AGE9"),
    ("DEMO.csv", "AGE60", "Copy", "Relation", "stg_Store", "AGE60"),
    ("DEMO.csv", "WORKWOM", "Copy", "Relation", "stg_Store", "WORKWOM"),
]

for row in data1:
    ws1.append(row)

for col in ws1.columns:
    ws1.column_dimensions[col[0].column_letter].width = 25

# --- MAPPING TABLE 2 ---
ws2 = wb.create_sheet("Mapping Table 2")

headers2 = ["Staging Table", "Staging Attribute", "Mapping", "DM Table Type", "DM Table", "DM Attribute"]
ws2.append(headers2)

for cell in ws2[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

data2 = [
    # stg_Movement -> FactWeeklySales
    ("stg_Movement_SDR/CSO/TPA/CRA", "MOVE", "Copy", "Fact", "FactWeeklySales", "units_sold"),
    ("stg_Movement_SDR/CSO/TPA/CRA", "PRICE", "Copy", "Fact", "FactWeeklySales", "shelf_price"),
    ("stg_Movement_SDR/CSO/TPA/CRA", "QTY", "Copy", "Fact", "FactWeeklySales", "price_qty"),
    ("stg_Movement_SDR/CSO/TPA/CRA", "PRICE, QTY", "Transform: PRICE/QTY", "Fact", "FactWeeklySales", "unit_price"),
    ("stg_Movement_SDR/CSO/TPA/CRA", "MOVE, PRICE, QTY", "Transform: MOVE*(PRICE/QTY)", "Fact", "FactWeeklySales", "revenue"),
    ("stg_Movement_SDR/CSO/TPA/CRA", "PROFIT", "Copy", "Fact", "FactWeeklySales", "gross_profit"),
    ("stg_Movement_SDR/CSO/TPA/CRA", "PROFIT, revenue", "Transform: (PROFIT/revenue)*100", "Fact", "FactWeeklySales", "profit_margin_pct"),
    ("stg_Movement_SDR/CSO/TPA/CRA", "UPC", "Lookup -> DimProduct", "Fact", "FactWeeklySales", "product_key"),
    ("stg_Movement_SDR/CSO/TPA/CRA", "STORE", "Lookup -> DimStore", "Fact", "FactWeeklySales", "store_key"),
    ("stg_Movement_SDR/CSO/TPA/CRA", "WEEK", "Lookup -> DimTime", "Fact", "FactWeeklySales", "time_key"),
    ("stg_Movement_SDR/CSO/TPA/CRA", "CATEGORY_CODE", "Lookup -> DimCategory", "Fact", "FactWeeklySales", "category_key"),
    ("stg_Movement_SDR/CSO/TPA/CRA", "SALE", "Lookup -> DimPromotion", "Fact", "FactWeeklySales", "promotion_key"),

    # stg_Product -> DimProduct
    ("stg_Product_SDR/CSO/TPA/CRA", "COM_CODE", "Copy", "Dimension", "DimProduct", "commodity_code"),
    ("stg_Product_SDR/CSO/TPA/CRA", "UPC", "Copy", "Dimension", "DimProduct", "upc"),
    ("stg_Product_SDR/CSO/TPA/CRA", "DESCRIP", "Transform: strip #, ~", "Dimension", "DimProduct", "description"),
    ("stg_Product_SDR/CSO/TPA/CRA", "SIZE", "Copy", "Dimension", "DimProduct", "size"),
    ("stg_Product_SDR/CSO/TPA/CRA", "CASE_PACK", "Copy", "Dimension", "DimProduct", "case_pack"),
    ("stg_Product_SDR/CSO/TPA/CRA", "NITEM", "Copy", "Dimension", "DimProduct", "item_number"),

    # stg_Store -> DimStore
    ("stg_Store", "STORE", "Copy (CAST to INT)", "Dimension", "DimStore", "store_id"),
    ("stg_Store", "NAME", "Transform: ISNULL -> 'UNKNOWN'", "Dimension", "DimStore", "store_name"),
    ("stg_Store", "CITY", "Transform: ISNULL -> 'UNKNOWN'", "Dimension", "DimStore", "city"),
    ("stg_Store", "ZIP", "Copy (CAST to INT)", "Dimension", "DimStore", "zip_code"),
    ("stg_Store", "ZONE", "Copy (CAST to INT)", "Dimension", "DimStore", "zone"),
    ("stg_Store", "WEEKVOL", "Copy (CAST to INT)", "Dimension", "DimStore", "weekly_volume"),
    ("stg_Store", "URBAN", "Copy (CAST to BIT)", "Dimension", "DimStore", "is_urban"),
    ("stg_Store", "INCOME", "Copy (CAST to DECIMAL)", "Dimension", "DimStore", "avg_income"),
    ("stg_Store", "EDUC", "Copy (CAST to DECIMAL)", "Dimension", "DimStore", "education_pct"),
    ("stg_Store", "POVERTY", "Copy (CAST to DECIMAL)", "Dimension", "DimStore", "poverty_pct"),
    ("stg_Store", "HSIZEAVG", "Copy (CAST to DECIMAL)", "Dimension", "DimStore", "avg_household_size"),
    ("stg_Store", "ETHNIC", "Copy (CAST to DECIMAL)", "Dimension", "DimStore", "ethnic_diversity"),
    ("stg_Store", "DENSITY", "Copy (CAST to DECIMAL)", "Dimension", "DimStore", "population_density"),
    ("stg_Store", "AGE9", "Copy (CAST to DECIMAL)", "Dimension", "DimStore", "age_under_9_pct"),
    ("stg_Store", "AGE60", "Copy (CAST to DECIMAL)", "Dimension", "DimStore", "age_over_60_pct"),
    ("stg_Store", "WORKWOM", "Copy (CAST to DECIMAL)", "Dimension", "DimStore", "working_women_pct"),
    ("stg_Store", "PRICLOW, PRICMED, PRICHIGH", "Transform: CASE WHEN", "Dimension", "DimStore", "price_tier"),

    # Generated/Hardcoded
    ("(generated)", "week_id 1-400", "Transform: DATEADD", "Dimension", "DimTime", "week_start_date, month, quarter..."),
    ("(hardcoded)", "28 category codes", "Direct INSERT", "Dimension", "DimCategory", "category_code, category_name..."),
    ("(hardcoded)", "4 deal types", "Direct INSERT", "Dimension", "DimPromotion", "deal_code, deal_type, is_promoted"),
]

for row in data2:
    ws2.append(row)

for col in ws2.columns:
    ws2.column_dimensions[col[0].column_letter].width = 25

wb.save("MappingTables.xlsx")
